"""SIT reference loading and detection-name resolution.

Two name sources feed dim_sit:

- The SIT risk workbook (the curated 18-column reference contract). The
  loader ports the legacy fork's GUID/slug metadata merge from
  build_activity_explorer_old_powerbi_data.load_sit_reference: Purview only
  reports GUIDs in detections, so custom slug-row metadata (Source, QGISCF,
  risk ratings, ...) is overlaid onto the GUID row with the same SIT name,
  and slug rows that duplicate a GUID name are dropped.
- An optional tenant GUID->name map (CurrentTenantSITs.json — the flat map
  this tool auto-generates on SIT export). It names observed SITs the
  workbook has no GUID row for, and bridges detections onto workbook
  slug rows when the names match.

``resolve_detected_sit`` is the per-detection resolution chain:
workbook GUID row > raw-payload name > tenant map > GUID fallback.
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from .errors import EnrichmentError

GUID_RE = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
)

# Workbook header -> dim_sit column (the legacy 18-column reference contract).
_WORKBOOK_COLUMNS = {
    "SIT Name": "sit_name",
    "GUID / Slug": "identifier",
    "Category": "category",
    "Risk Description": "risk_description",
    "Risk Rating (1-10)": "risk_score",
    "Reference URL": "reference_url",
    "Australian PSPF Classification": "pspf_classification",
    "QGISCF": "qgiscf",
    "QGISCF DLM": "qgiscf_dlm",
    "Label Code": "label_code",
    "Classifier Type": "sit_classifier_type",
    "Source": "source",
    "Jurisdictions": "jurisdictions",
    "Scope": "scope",
    "Confidence": "reference_confidence",
    "Classification Tier": "classification_tier",
    "Generic Classification": "generic_classification",
    "Generic DLM": "generic_dlm",
    # Extras the v5 loader carried from the workbook / cross-reference sheet.
    "Data Categories": "data_categories",
    "Regulations": "regulations",
    "Small (tenant)": "small_tenant",
    "Medium (tenant)": "medium_tenant",
    "Large (tenant)": "large_tenant",
}

_BOOL_COLUMNS = {"small_tenant", "medium_tenant", "large_tenant"}

_RISK_SHEET = "SIT Risk Analysis"

_RISK_GLOB = "SIT-Risk-Analysis*.xlsx"

_SIT_NAMES_FILENAME = "CurrentTenantSITs.json"


@dataclass
class RiskLookup:
    rows: list[dict[str, Any]] = field(default_factory=list)
    by_name: dict[str, dict[str, Any]] = field(default_factory=dict)
    by_id: dict[str, dict[str, Any]] = field(default_factory=dict)
    by_key: dict[str, dict[str, Any]] = field(default_factory=dict)
    source_path: Path | None = None


def _norm_text(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip()).lower()


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _cell_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    return str(value).strip().upper() in {"Y", "YES", "TRUE", "1"}


def risk_band(score: int | None) -> str:
    if score is None:
        return "Unrated"
    if score >= 9:
        return "Critical"
    if score >= 7:
        return "High"
    if score >= 4:
        return "Medium"
    return "Low"


def sit_key_for(name: str | None, identifier: str | None) -> str:
    identifier = (identifier or "").strip()
    if identifier:
        return identifier.lower()
    return f"name:{_norm_text(name)}"


def _read_worksheet_rows(ws) -> Iterable[dict[str, Any]]:
    header: list[str] | None = None
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if header is None:
            header = [str(v).strip() if v is not None else "" for v in values]
            continue
        if not any(v is not None and str(v).strip() for v in values):
            continue
        yield {
            header[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(header))
            if header[idx]
        }


def _workbook_row(raw: dict[str, Any]) -> dict[str, Any] | None:
    """Map one worksheet row to dim_sit-shaped fields (identifier kept aside)."""
    row: dict[str, Any] = {}
    for header, column in _WORKBOOK_COLUMNS.items():
        value = raw.get(header)
        if column == "risk_score":
            row[column] = _cell_int(value)
        elif column in _BOOL_COLUMNS:
            row[column] = _cell_bool(value)
        else:
            row[column] = _cell_str(value)
    if not row.get("sit_name"):
        return None
    return row


def _merge_guid_slug_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Port of the legacy fork's GUID/slug metadata merge (keyed by SIT name).

    Name matching is exact (case-sensitive) to reproduce the legacy
    sit_reference row set byte-for-byte: rows whose names differ only by case
    are distinct SIT entries there.
    """
    guid_rows = [r for r in rows if GUID_RE.match(r.get("identifier") or "")]
    slug_rows = [r for r in rows if not GUID_RE.match(r.get("identifier") or "")]

    slug_by_name: dict[str, dict[str, Any]] = {}
    for row in slug_rows:
        slug_by_name.setdefault(row["sit_name"], row)

    for row in guid_rows:
        custom = slug_by_name.get(row["sit_name"])
        if not custom:
            continue
        for column, value in custom.items():
            if column in ("identifier", "sit_name"):
                continue
            if value is not None and str(value).strip():
                row[column] = value

    guid_names = {r["sit_name"] for r in guid_rows}
    orphan_slugs = [r for r in slug_rows if r["sit_name"] not in guid_names]
    return [*guid_rows, *orphan_slugs]


def _finalize_row(row: dict[str, Any], source_sheet: str) -> dict[str, Any]:
    identifier = (row.pop("identifier", None) or "").strip()
    is_guid = bool(GUID_RE.match(identifier))
    score = row.get("risk_score")
    row.update(
        sit_key=sit_key_for(row.get("sit_name"), identifier),
        sit_id=identifier.lower() if is_guid else None,
        sit_slug=identifier.lower() if identifier and not is_guid else None,
        risk_band=risk_band(score),
        source_sheet=source_sheet,
        is_unrated=score is None,
    )
    return row


def load_risk_workbook(path: Path) -> RiskLookup:
    """Load the SIT risk workbook into dim_sit-shaped rows plus lookups."""
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - environment issue
        raise EnrichmentError(
            f"Missing Python dependency '{exc.name}'. "
            "Install runtime dependencies with `pip install -r requirements.txt`."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if _RISK_SHEET in workbook.sheetnames:
            sheet = workbook[_RISK_SHEET]
            sheet_name = _RISK_SHEET
        else:
            sheet = workbook[workbook.sheetnames[0]]
            sheet_name = workbook.sheetnames[0]
        raw_rows = [
            row for raw in _read_worksheet_rows(sheet)
            if (row := _workbook_row(raw)) is not None
        ]
    finally:
        workbook.close()

    merged = _merge_guid_slug_rows(raw_rows)
    rows = [_finalize_row(row, sheet_name) for row in merged]

    lookup = RiskLookup(rows=rows, source_path=path)
    for row in rows:
        lookup.by_key.setdefault(row["sit_key"], row)
        name_norm = _norm_text(row["sit_name"])
        if name_norm:
            lookup.by_name.setdefault(name_norm, row)
        if row.get("sit_id"):
            lookup.by_id.setdefault(row["sit_id"], row)
    return lookup


def sit_key_for_detected_id(sit_id: str | None, risk: RiskLookup) -> str:
    """Resolve a detection's SensitiveInfoTypeId to a dim_sit key."""
    sit_id_norm = (sit_id or "").strip().lower()
    row = risk.by_id.get(sit_id_norm)
    if row:
        return row["sit_key"]
    return sit_id_norm or "unknown"


def load_sit_name_map(path: Path) -> dict[str, str]:
    """GUID (lowercase) -> display name from a CurrentTenantSITs.json-style file.

    The expected shape is the flat ``{"<guid>": "<name>"}`` map this tool
    auto-generates on SIT export (ConfigFiles/CurrentTenantSITs.json);
    properties starting with ``_`` are metadata and non-GUID keys are
    ignored. Malformed JSON or a non-object payload is a hard error — a name
    map that cannot be read must never silently degrade to GUID labels.
    """
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise EnrichmentError(f"Cannot read SIT name map {path}: {exc}") from exc
    if not isinstance(payload, dict):
        raise EnrichmentError(f"SIT name map is not a JSON object: {path}")
    names: dict[str, str] = {}
    for key, value in payload.items():
        key = str(key).strip()
        if key.startswith("_") or not GUID_RE.match(key):
            continue
        text = _cell_str(value)
        if text:
            names[key.lower()] = text
    return names


def _search_one(input_dir: Path, pattern: str) -> Path | None:
    """Search the export root, then exactly one directory level below it."""
    direct = sorted(input_dir.glob(pattern))
    if direct:
        return direct[-1]
    one_level = sorted(input_dir.glob(f"*/{pattern}"))
    if one_level:
        return one_level[-1]
    return None


def resolve_risk_workbook(input_dir: Path, explicit: Path | None) -> Path | None:
    if explicit is not None:
        if not explicit.exists():
            raise EnrichmentError(f"--risk-workbook does not exist: {explicit}")
        return explicit
    return _search_one(input_dir, _RISK_GLOB)


def resolve_sit_names_path(
    input_dir: Path, explicit: Path | None, default: Path | None = None,
) -> Path | None:
    """SIT name-map resolution: --sit-names > export root (one level deep) >
    repo-default ConfigFiles/CurrentTenantSITs.json. Names are best-effort,
    so a missing map is fine (GUID fallback) — but an explicit path that does
    not exist is a hard error."""
    if explicit is not None:
        if not explicit.exists():
            raise EnrichmentError(f"--sit-names does not exist: {explicit}")
        return explicit
    found = _search_one(input_dir, _SIT_NAMES_FILENAME)
    if found is not None:
        return found
    if default is not None and default.exists():
        return default
    return None


def resolve_detected_sit(
    sit_id: str | None,
    raw_name: str | None,
    risk: RiskLookup,
    tenant_names: dict[str, str] | None = None,
) -> tuple[str, dict[str, Any] | None, str | None, str | None, bool]:
    """Resolve one detection to (sit_key, workbook_row, name_source,
    display_name, bridged_to_workbook).

    Display-name resolution chain:
      workbook GUID row > raw-payload name > tenant GUID->name map > GUID.
    The workbook wins for GUIDs it knows: its names are curated (it is itself
    derived from a tenant SIT export) and the exclusion list keys on them.
    When only a name source matches (raw payload / tenant map) and the
    workbook has a row with that NAME — custom SITs live in the workbook as
    slug rows while detections only ever carry GUIDs — the detection is
    bridged onto the workbook row so its metadata, risk score and exclusion
    behavior apply and no duplicate display-name dim row is created.
    """
    sit_id_norm = (sit_id or "").strip().lower()
    row = risk.by_id.get(sit_id_norm)
    if row is not None:
        return row["sit_key"], row, "workbook", row["sit_name"], False
    fallback_key = sit_id_norm or "unknown"
    for name, source in ((raw_name, "raw_payload"),
                         ((tenant_names or {}).get(sit_id_norm), "tenant_map")):
        if not name:
            continue
        bridged = risk.by_name.get(_norm_text(name))
        if bridged is not None:
            return bridged["sit_key"], bridged, source, bridged["sit_name"], True
        return fallback_key, None, source, name, False
    return fallback_key, None, None, None, False
