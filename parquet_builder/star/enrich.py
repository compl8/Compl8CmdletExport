"""Enrichment input loaders: SIT risk workbook and GAL/department mapping.

F1 fix policy: enrichment inputs are resolved from explicit CLI paths first,
then the export root, then one directory level below it. If either input is
still missing the conversion FAILS unless --allow-unenriched was passed —
silent unenriched output (risk scores all zero, one department) poisoned the
v5 run and is no longer possible.

The workbook loader ports the legacy fork's GUID/slug metadata merge from
build_activity_explorer_old_powerbi_data.load_sit_reference: Purview only
reports GUIDs in detections, so custom slug-row metadata (Source, QGISCF,
risk ratings, ...) is overlaid onto the GUID row with the same SIT name, and
slug rows that duplicate a GUID name are dropped. All reference columns the
legacy sit_reference carried are parsed (plus the cross-reference extras the
v5 loader knew about).
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

GUID_RE = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
)

# Workbook header -> dim_sit column (the legacy 18-column reference contract).
_WORKBOOK_COLUMNS = {
    "SIT Name": "sit_name",
    "GUID / Slug": "identifier",
    "Category": "category",
    "Risk Description": "risk_description",
    "Risk Rating (1-10)": "risk_score",
    "Reference URL": "reference_url",
    "Australian PSPF Classification": "pspf_classification",
    "QGISCF": "qgiscf",
    "QGISCF DLM": "qgiscf_dlm",
    "Label Code": "label_code",
    "Classifier Type": "sit_classifier_type",
    "Source": "source",
    "Jurisdictions": "jurisdictions",
    "Scope": "scope",
    "Confidence": "reference_confidence",
    "Classification Tier": "classification_tier",
    "Generic Classification": "generic_classification",
    "Generic DLM": "generic_dlm",
    # Extras the v5 loader carried from the workbook / cross-reference sheet.
    "Data Categories": "data_categories",
    "Regulations": "regulations",
    "Small (tenant)": "small_tenant",
    "Medium (tenant)": "medium_tenant",
    "Large (tenant)": "large_tenant",
}

_BOOL_COLUMNS = {"small_tenant", "medium_tenant", "large_tenant"}

_RISK_SHEET = "SIT Risk Analysis"

_DEPARTMENT_FILE_CANDIDATES = (
    "user_department_mapping.csv",
    "user_department_mapping.xlsx",
    "User-Department-Mapping.csv",
    "User-Department-Mapping.xlsx",
    "GAL_Clean.csv",
)

_RISK_GLOB = "SIT-Risk-Analysis*.xlsx"


class EnrichmentError(RuntimeError):
    """Raised when required enrichment inputs cannot be resolved."""


@dataclass
class RiskLookup:
    rows: list[dict[str, Any]] = field(default_factory=list)
    by_name: dict[str, dict[str, Any]] = field(default_factory=dict)
    by_id: dict[str, dict[str, Any]] = field(default_factory=dict)
    by_key: dict[str, dict[str, Any]] = field(default_factory=dict)
    source_path: Path | None = None


def _norm_text(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip()).lower()


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _cell_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    return str(value).strip().upper() in {"Y", "YES", "TRUE", "1"}


def risk_band(score: int | None) -> str:
    if score is None:
        return "Unrated"
    if score >= 9:
        return "Critical"
    if score >= 7:
        return "High"
    if score >= 4:
        return "Medium"
    return "Low"


def sit_key_for(name: str | None, identifier: str | None) -> str:
    identifier = (identifier or "").strip()
    if identifier:
        return identifier.lower()
    return f"name:{_norm_text(name)}"


def _read_worksheet_rows(ws) -> Iterable[dict[str, Any]]:
    header: list[str] | None = None
    for row in ws.iter_rows(values_only=True):
        values = list(row)
        if header is None:
            header = [str(v).strip() if v is not None else "" for v in values]
            continue
        if not any(v is not None and str(v).strip() for v in values):
            continue
        yield {
            header[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(header))
            if header[idx]
        }


def _workbook_row(raw: dict[str, Any]) -> dict[str, Any] | None:
    """Map one worksheet row to dim_sit-shaped fields (identifier kept aside)."""
    row: dict[str, Any] = {}
    for header, column in _WORKBOOK_COLUMNS.items():
        value = raw.get(header)
        if column == "risk_score":
            row[column] = _cell_int(value)
        elif column in _BOOL_COLUMNS:
            row[column] = _cell_bool(value)
        else:
            row[column] = _cell_str(value)
    if not row.get("sit_name"):
        return None
    return row


def _merge_guid_slug_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Port of the legacy fork's GUID/slug metadata merge (keyed by SIT name).

    Name matching is exact (case-sensitive) to reproduce the legacy
    sit_reference row set byte-for-byte: rows whose names differ only by case
    are distinct SIT entries there.
    """
    guid_rows = [r for r in rows if GUID_RE.match(r.get("identifier") or "")]
    slug_rows = [r for r in rows if not GUID_RE.match(r.get("identifier") or "")]

    slug_by_name: dict[str, dict[str, Any]] = {}
    for row in slug_rows:
        slug_by_name.setdefault(row["sit_name"], row)

    for row in guid_rows:
        custom = slug_by_name.get(row["sit_name"])
        if not custom:
            continue
        for column, value in custom.items():
            if column in ("identifier", "sit_name"):
                continue
            if value is not None and str(value).strip():
                row[column] = value

    guid_names = {r["sit_name"] for r in guid_rows}
    orphan_slugs = [r for r in slug_rows if r["sit_name"] not in guid_names]
    return [*guid_rows, *orphan_slugs]


def _finalize_row(row: dict[str, Any], source_sheet: str) -> dict[str, Any]:
    identifier = (row.pop("identifier", None) or "").strip()
    is_guid = bool(GUID_RE.match(identifier))
    score = row.get("risk_score")
    row.update(
        sit_key=sit_key_for(row.get("sit_name"), identifier),
        sit_id=identifier.lower() if is_guid else None,
        sit_slug=identifier.lower() if identifier and not is_guid else None,
        risk_band=risk_band(score),
        source_sheet=source_sheet,
        is_unrated=score is None,
    )
    return row


def load_risk_workbook(path: Path) -> RiskLookup:
    """Load the SIT risk workbook into dim_sit-shaped rows plus lookups."""
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - environment issue
        raise EnrichmentError(
            f"Missing Python dependency '{exc.name}'. "
            "Install runtime dependencies with `pip install -r requirements.txt`."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if _RISK_SHEET in workbook.sheetnames:
            sheet = workbook[_RISK_SHEET]
            sheet_name = _RISK_SHEET
        else:
            sheet = workbook[workbook.sheetnames[0]]
            sheet_name = workbook.sheetnames[0]
        raw_rows = [
            row for raw in _read_worksheet_rows(sheet)
            if (row := _workbook_row(raw)) is not None
        ]
    finally:
        workbook.close()

    merged = _merge_guid_slug_rows(raw_rows)
    rows = [_finalize_row(row, sheet_name) for row in merged]

    lookup = RiskLookup(rows=rows, source_path=path)
    for row in rows:
        lookup.by_key.setdefault(row["sit_key"], row)
        name_norm = _norm_text(row["sit_name"])
        if name_norm:
            lookup.by_name.setdefault(name_norm, row)
        if row.get("sit_id"):
            lookup.by_id.setdefault(row["sit_id"], row)
    return lookup


def sit_key_for_detected_id(sit_id: str | None, risk: RiskLookup) -> str:
    """Resolve a detection's SensitiveInfoTypeId to a dim_sit key."""
    sit_id_norm = (sit_id or "").strip().lower()
    row = risk.by_id.get(sit_id_norm)
    if row:
        return row["sit_key"]
    return sit_id_norm or "unknown"


def _mapping_value(row: dict[str, Any], names: tuple[str, ...]) -> str | None:
    normalized = {re.sub(r"[^a-z0-9]+", "", key.lower()): value for key, value in row.items()}
    for name in names:
        value = normalized.get(re.sub(r"[^a-z0-9]+", "", name.lower()))
        if value is not None and str(value).strip():
            return str(value).strip()
    return None


def _read_mapping_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(_read_worksheet_rows(workbook.active))
        finally:
            workbook.close()
    return []


def load_department_mapping(path: Path) -> dict[str, dict[str, Any]]:
    """User key (uppercased) -> department mapping rows from a GAL/department file.

    Each row is registered under BOTH its UserPrincipalName and its Mail
    address: Activity Explorer identifies users by primary SMTP address,
    which routinely lives on a different domain than the UPN (e.g. UPN
    user@qfes.qld.gov.au vs mail user@fire.qld.gov.au). Keying on UPN alone
    silently drops those users into 'Unmapped' at fact grain. UPN keys win
    collisions; mail-derived entries carry ``is_alias`` so dim_user seeding
    can skip them (one dim_user row per person, not per address).

    Department names are case-canonicalized across the file (most frequent
    casing wins) so variants like 'qfes'/'QFES' cannot split into two
    dim_department rows — Power BI's case-insensitive engine merges such
    labels and displays an arbitrary casing.
    """
    parsed: list[tuple[str | None, str | None, dict[str, Any]]] = []
    casing_counts: dict[str, dict[str, int]] = {}
    for row in _read_mapping_rows(path):
        upn = _mapping_value(row, ("userprincipalname", "user_upn", "upn", "user", "username"))
        mail = _mapping_value(row, ("mail", "email", "primarysmtpaddress"))
        if not upn and not mail:
            continue
        department = _mapping_value(row, ("department", "dept")) or "Unmapped"
        counts = casing_counts.setdefault(department.casefold(), {})
        counts[department] = counts.get(department, 0) + 1
        parsed.append((upn, mail, {
            "department": department,
            "division": _mapping_value(row, ("division", "directorate", "group")),
            "business_unit": _mapping_value(
                row, ("business_unit", "businessunit", "unit", "section", "branch", "team")
            ),
            "mapping_source": path.name,
            "is_mapped": True,
        }))

    canonical_casing = {
        folded: max(counts.items(), key=lambda item: item[1])[0]
        for folded, counts in casing_counts.items()
    }

    mappings: dict[str, dict[str, Any]] = {}
    for upn, mail, entry in parsed:
        entry["department"] = canonical_casing[entry["department"].casefold()]
        primary = upn or mail
        mappings[primary.upper().strip()] = entry
    # Mail aliases never displace a primary (UPN) key.
    for upn, mail, entry in parsed:
        if not (upn and mail):
            continue
        alias_key = mail.upper().strip()
        if alias_key and alias_key not in mappings:
            mappings[alias_key] = {**entry, "is_alias": True}
    return mappings


def _search_one(input_dir: Path, pattern: str) -> Path | None:
    """Search the export root, then exactly one directory level below it."""
    direct = sorted(input_dir.glob(pattern))
    if direct:
        return direct[-1]
    one_level = sorted(input_dir.glob(f"*/{pattern}"))
    if one_level:
        return one_level[-1]
    return None


def resolve_risk_workbook(input_dir: Path, explicit: Path | None) -> Path | None:
    if explicit is not None:
        if not explicit.exists():
            raise EnrichmentError(f"--risk-workbook does not exist: {explicit}")
        return explicit
    return _search_one(input_dir, _RISK_GLOB)


def resolve_department_csv(input_dir: Path, explicit: Path | None) -> Path | None:
    if explicit is not None:
        if not explicit.exists():
            raise EnrichmentError(f"--department-csv does not exist: {explicit}")
        return explicit
    for name in _DEPARTMENT_FILE_CANDIDATES:
        found = _search_one(input_dir, name)
        if found is not None:
            return found
    return None


def resolve_enrichment_inputs(
    input_dir: Path,
    risk_workbook: Path | None,
    department_csv: Path | None,
    allow_unenriched: bool,
) -> tuple[Path | None, Path | None]:
    """Resolve both enrichment inputs, enforcing the F1 hard-fail policy."""
    risk_path = resolve_risk_workbook(input_dir, risk_workbook)
    dept_path = resolve_department_csv(input_dir, department_csv)

    missing = []
    if risk_path is None:
        missing.append(f"SIT risk workbook ({_RISK_GLOB})")
    if dept_path is None:
        missing.append(
            "department mapping (" + ", ".join(_DEPARTMENT_FILE_CANDIDATES) + ")"
        )
    if missing and not allow_unenriched:
        raise EnrichmentError(
            "Enrichment inputs not found: "
            + "; ".join(missing)
            + f". Searched {input_dir} and one level below. "
            "Pass --risk-workbook/--department-csv, or --allow-unenriched to "
            "deliberately produce an unenriched model (risk scores will be 0)."
        )
    return risk_path, dept_path
